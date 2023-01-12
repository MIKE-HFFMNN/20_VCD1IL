import matplotlib.pyplot as plt
import matplotlib as mpl
import numpy as np
import os
import math
from openpyxl import load_workbook
import tkinter as tk
from tkinter import *
import argparse

#Setup argparser
arg_parser_ = argparse.ArgumentParser(description="Process some integers.")

#Definition of basic variables and start values
g = 9.81
v = 0.0
vm = 0.0
m = 0.0
y = 0.0

#Definition of basic formulas from lecture
def calcsn(v):
    global sn
    sn = float(pow(v/10, 2))
    return sn
def calcsd(v):
    global sd
    sd = float(pow(v/10, 2)/2)
    return sd
def calcsrc(v):
    global src
    src = float((v/10)*3)
    return src
def calcst(src, sn):
    global st
    st = src+sn
    return st
"""
def calcfn(m, g):
    fn = float(m*g)
    return fn
"""
#Definition of derived and simplified formulas (https://austria-forum.org/af/AustriaWiki/Bremswe and https://www.w3schools.com/python/ref_math_sin.asp)
def calcamax(rs, g, y):
    amax = float(rs*g*math.cos(math.radians(y))+g*math.sin(math.radians(y)))
    return amax
def calctmax(rs, g, y, vm):
    amax = float(rs*g*math.cos(math.radians(y))+g*math.sin(math.radians(y)))
    return vm/amax

#User input velocity (https://realpython.com/python-gui-tkinter/)
root = Tk()
root.title('Breaking distance simulation')
root.geometry('258x215')
def inputv():
    global v
    global vm
    try:
        v = float(box.get("1.0", "end-1c"))
    except ValueError:
        answer.config(text="Please enter a number!")

    while True:
        if v >= 1 and v <= 300:
            print("Velocity input: ", v, " kp/h")
            vm = float(v/3.6)
            print("Velocity input: ", vm, " m/s")
            return vm, v
        else:
            error = tk.Label(text="Error - please enter velocity >0 and <300 kph!", fg="black", width=40, height=5)
            error.pack()
            print("Error - please enter velocity >0 and <300 kph!")
            mainloop()

label = tk.Label(text="Hello! Please insert velocity:",fg="black",width=30,height=5)
label.pack()
box = Text(root, height=1, width=8)
box.pack()
button = Button(root, height=1, width=16, text="Enter velocity in kp/h", command=lambda: inputv())
button.pack()
answer = Label(root, text=" ")
answer.pack(pady=20)

mainloop()

"""
#User input vehicle mass
root = Tk()
root.geometry('258x215')
def inputm():
    m = float(txt_Box.get("1.0", "end-1c"))
    print("Mass input: ", m, " kg")
txt_Box = Text(root, height= 1, width=8)
txt_Box.pack()
button = Button(root, height=1, width= 16, text="Enter mass in kg", command=lambda: inputm())
button.pack()
mainloop()
"""

#User input inclination
root = Tk()
root.title('Breaking distance simulation')
root.geometry('258x215')
def inputy():
    global y
    try:
        y = float(box.get("1.0", "end-1c"))
    except ValueError:
        answer.config(text="Please enter a number!")

    while True:
        if y >= 0 and y <= 45:
            print("Inclination input: ", y, "°")
            return y
        else:
            error = tk.Label(text="Error - please enter inclination <45°!", fg="black", width=30, height=5)
            error.pack()
            print("Error - please enter inclination <45°!")
            mainloop()

label = tk.Label(text="Please insert inclination angle:",fg="black",width=30,height=5)
label.pack()
box = Text(root, height=1, width=8)
box.pack()
button = Button(root, height=1, width=16, text="Enter inclination in °", command=lambda: inputy())
button.pack()
answer = Label(root, text=" ")
answer.pack(pady=20)
mainloop()

#User selection of road surface via dropdown (https://stackoverflow.com/questions/52757496/how-to-make-tkinter-drop-down-to-save-data-in-python-3)
root = tk.Tk()
root.title('Breaking distance simulation')
root.geometry('258x215')
servs = ['concrete', 'ice', 'water', 'gravel', 'sand']
svar = tk.StringVar()
svar.set(servs[0])
sr = servs[0]
def _get(cur):
    global sr
    sr = str(cur)
    print("Road surface input: ", sr)
    return sr
drop = tk.OptionMenu(root, svar, command = _get, *servs)
drop.grid(row=2, column=1)
root.mainloop()

#User selection of road condition via dropdown
root = tk.Tk()
root.title('Breaking distance simulation')
root.geometry('258x215')
servs = ['dry', 'wet', 'aquaplaning']
svar = tk.StringVar()
svar.set(servs[0])
sc = str(servs[0])
def _get(cur):
    global sc
    sc = str(cur)
    print("Road condition input: ", sc)
    return sc
drop = tk.OptionMenu(root, svar, command = _get, *servs)
drop.grid(row=2, column=1)
root.mainloop()

#Call and calculate distance s according rule of thumb formula from class for comparison
sn = calcsn(v)
print("Rule of thumb distance (normal) :", sn, "m")
sd = calcsd(v)
print("Rule of thumb distance (danger) :", sd, "m")
src = calcsrc(v)
print("Rule of thumb distance (reaction) :", src, "m")
st = calcst(src, sn)
print("Rule of thumb distance (normal + reaction) :", st, "m")

#Access Excel file for corresponding surface road (sr) friction values
wb = load_workbook('20_VCD1IL_CODING.xlsx', data_only=True)
print(wb.sheetnames)
ws = wb["Friction values"]
ms = [cell.value for cell in ws['C'][1:]]
md = [cell.value for cell in ws['D'][1:]]
print(ms)
print(md)

#Defintion of distance s calculation depending on µs from Excel file
def calcs(sr, sc, y, vm):
    while True:
        if sr == "concrete" and sc == "dry":
            rs = float(ms[0])
            rd = float(md[0])
        elif sr == "concrete" and sc == "wet":
            rs = float(ms[1])
            rd = float(md[1])
        elif sr == "ice" and sc == "dry":
            rs = float(ms[2])
            rd = float(md[2])
        elif sr == "ice" and sc == "wet":
            rs = float(ms[3])
            rd = float(md[3])
        elif sr == "water" and sc == "aquaplaning":
            rs = float(ms[4])
            rd = float(md[4])
        elif sr == "gravel" and sc == "dry":
            rs = float(ms[5])
            rd = float(md[5])
        elif sr == "sand" and sc == "dry":
            rs = float(ms[6])
            rd = float(md[6])
        else:
            error = tk.Label(text="Error - option not possible!", fg="black", width=30, height=5)
            error.pack()
            print("Error - option not possible!")
            mainloop()
            exit()
        print("rs: ", rs)
        print("rd: ", rd)

        #Call calculation of tmax, trounded and creation of equally sized vectors for plotting
        global tmax
        tmax = calctmax(rs, g, y, vm)
        print("tmax :", tmax, "s")
        amax = calcamax(rs, g, y)
        print("amax :", amax, "m/s^2")
        tr = abs(math.ceil(tmax))
        print("tmax rounded-up :", tr, "s")
        tv = np.arange(0, tr, 0.1)
        sv = [None] * len(tv)
        vv = [None] * len(tv)
        vv[0] = vm

        #Creation of distance vector acc. formula from lecture, velocity vector until standstill at tmax for plotting, and reference vector for rule of thumb distance comparison
        i = 0
        while i < len(tv):
            ti = tv[1] - tv[0]
            sv[i] = vm * tv[i] - 1/2 * calcamax(rs, g, y) * pow(tv[i], 2)
            #vv[i + 1] = vv[i] - calcamax(rs, g, y) * ti
            if vv[i] <= 0:
                break
            vv[i + 1] = vv[i] - calcamax(rs, g, y) * ti
            i += 1
        print(tv, "[s]")
        print(sv, "[m]")
        print(vv, "[m/s]")
        return tv, sv, vv

#Call calculation of s by creating 3D array of tv, sv and vv
array = calcs(sr,sc,y,vm)
ta = array[0]
sa = array[1]
va = array[2]
print(array)

#Plotting of velocity v and distance s over time t (https://stackoverflow.com/questions/14762181/adding-a-y-axis-label-to-secondary-y-axis-in-matplotlib)
#Twin axis (https://stackoverflow.com/questions/5484922/secondary-axis-with-twinx-how-to-add-to-legend)
#Start from 0 (https://stackoverflow.com/questions/22642511/change-y-range-to-start-from-0-with-matplotlib)
fig, ax1 = plt.subplots()
ax2 = ax1.twinx()

ax1.plot(ta, va, '-', label='velocity v', color='b')
ax2.plot(ta, sa, '-', label='distance s', color='r')
ax2.axhline(sn, label='reference distance sr', color='y', linestyle='dashed')

ax1.set_xlabel('time [s]')
ax1.set_ylabel('velocity [m/s]')
ax1.set_ylim(bottom=0)
ax1.set_xlim(left=0)
ax1.grid()

ax2.set_ylabel('distance [m]')
ax2.set_ylim(bottom=0)

lines1, labels1 = ax1.get_legend_handles_labels()
lines2, labels2 = ax2.get_legend_handles_labels()
ax2.legend(lines1 + lines2, labels1 + labels2, loc=0)

plt.title("velocity and distance over time")
plt.show()

#Comparison with online calculator (https://www.johannes-strommer.com/rechner/bremsweg-beschleunigung-geschwindigkeit/)